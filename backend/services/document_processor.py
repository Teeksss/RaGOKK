# Last reviewed: 2025-04-30 05:56:23 UTC (User: Teeksss)
from typing import Dict, Any, List, Optional
import os
import magic
import re
import logging
from datetime import datetime
import tempfile
import shutil

# PDF işleme
from pypdf import PdfReader

# İmaj işleme ve OCR için
try:
    from PIL import Image
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# Diğer dosya türleri
import docx2txt
try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from bs4 import BeautifulSoup
    HAS_HTML_PARSER = True
except ImportError:
    HAS_HTML_PARSER = False

logger = logging.getLogger(__name__)

class DocumentProcessorService:
    """
    Belge işleme ve metin çıkarma servisi
    """
    
    def __init__(self):
        """Servis başlangıç ayarları"""
        # MIME tip eşleştirmeleri
        self.mime_types = {
            'application/pdf': 'pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'docx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'xlsx',
            'text/plain': 'txt',
            'text/html': 'html',
            'text/markdown': 'md',
            'image/jpeg': 'jpg',
            'image/png': 'png',
            'image/tiff': 'tiff',
            'image/webp': 'webp'
        }
        
        # OCR aktif mi kontrol et
        self.has_ocr = HAS_OCR
    
    async def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Dosya hakkında bilgi döndürür
        
        Args:
            file_path: Dosya yolu
            
        Returns:
            Dict[str, Any]: Dosya bilgileri
        """
        try:
            # Dosya boyutu
            file_size = os.path.getsize(file_path)
            
            # MIME türü
            mime = magic.Magic(mime=True)
            mime_type = mime.from_file(file_path)
            
            # Dosya tipi tanımla
            file_type = self.mime_types.get(mime_type, 'unknown')
            
            # Temel dosya ismi
            base_name = os.path.basename(file_path)
            
            return {
                "file_path": file_path,
                "file_name": base_name,
                "file_size": file_size,
                "mime_type": mime_type,
                "file_type": file_type,
                "modified_time": datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat(),
                "created_time": datetime.fromtimestamp(os.path.getctime(file_path)).isoformat()
            }
        except Exception as e:
            logger.error(f"Error getting file info: {str(e)}")
            return {
                "file_path": file_path,
                "error": str(e),
                "file_type": "unknown"
            }
    
    async def extract_basic_text(self, file_path: str, file_type: Optional[str] = None) -> str:
        """
        Dosyadan basit metin çıkarma (OCR olmadan)
        
        Args:
            file_path: Dosya yolu
            file_type: Dosya türü (belirtilmezse otomatik algılanır)
            
        Returns:
            str: Çıkarılan metin
        """
        try:
            # Dosya türü belirtilmemişse algıla
            if not file_type:
                file_info = await self.get_file_info(file_path)
                file_type = file_info.get("file_type", "unknown")
            
            # Dosya türüne göre metin çıkarma
            if file_type == 'pdf':
                return await self._extract_text_from_pdf(file_path, use_ocr=False)
            elif file_type == 'docx':
                return await self._extract_text_from_docx(file_path)
            elif file_type == 'txt':
                return await self._extract_text_from_text(file_path)
            elif file_type == 'html':
                return await self._extract_text_from_html(file_path)
            elif file_type == 'md':
                return await self._extract_text_from_text(file_path)
            elif file_type == 'xlsx' and HAS_EXCEL:
                return await self._extract_text_from_excel(file_path)
            elif file_type in ['jpg', 'png', 'tiff', 'webp']:
                # Görüntü dosyaları için basit bir mesaj döndür
                return f"[Image File: OCR processing is required to extract text from this {file_type.upper()} image]"
            else:
                return f"[Unsupported file format: {file_type}]"
                
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            return f"[Error extracting text: {str(e)}]"
    
    async def extract_text_from_file(self, file_path: str, file_type: Optional[str] = None) -> str:
        """
        Dosyadan tam metin çıkarma (OCR dahil)
        
        Args:
            file_path: Dosya yolu
            file_type: Dosya türü (belirtilmezse otomatik algılanır)
            
        Returns:
            str: Çıkarılan metin
        """
        try:
            # Dosya türü belirtilmemişse algıla
            if not file_type:
                file_info = await self.get_file_info(file_path)
                file_type = file_info.get("file_type", "unknown")
            
            # Dosya türüne göre metin çıkarma
            if file_type == 'pdf':
                return await self._extract_text_from_pdf(file_path, use_ocr=True)
            elif file_type == 'docx':
                return await self._extract_text_from_docx(file_path)
            elif file_type == 'txt':
                return await self._extract_text_from_text(file_path)
            elif file_type == 'html':
                return await self._extract_text_from_html(file_path)
            elif file_type == 'md':
                return await self._extract_text_from_text(file_path)
            elif file_type == 'xlsx' and HAS_EXCEL:
                return await self._extract_text_from_excel(file_path)
            elif file_type in ['jpg', 'png', 'tiff', 'webp'] and self.has_ocr:
                return await self._extract_text_from_image(file_path)
            else:
                return f"[Unsupported file format: {file_type}]"
                
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            return f"[Error extracting text: {str(e)}]"
    
    async def _extract_text_from_pdf(self, file_path: str, use_ocr: bool = False) -> str:
        """PDF'den metin çıkarır"""
        try:
            # PDF okuyucuyu başlat
            reader = PdfReader(file_path)
            text_content = []
            
            # OCR desteği varsa ve OCR kullanılacaksa
            if use_ocr and self.has_ocr:
                # Her sayfa için
                for page_num in range(len(reader.pages)):
                    # Sayfadan metin çıkarmayı dene
                    page = reader.pages[page_num]
                    page_text = page.extract_text()
                    
                    # Sayfada metin yoksa veya çok azsa OCR uygula
                    if not page_text or len(page_text.strip()) < 100:
                        # PDF sayfasını geçici görüntüye dönüştür ve OCR uygula
                        image_text = await self._apply_ocr_to_pdf_page(file_path, page_num)
                        if image_text:
                            page_text = image_text
                    
                    # Sayfa başlığı ve metni ekle
                    text_content.append(f"\n--- Page {page_num + 1} ---\n")
                    text_content.append(page_text)
            else:
                # Basit metin çıkarma (OCR olmadan)
                for page_num, page in enumerate(reader.pages):
                    text_content.append(f"\n--- Page {page_num + 1} ---\n")
                    text_content.append(page.extract_text())
            
            return "\n".join(text_content)
            
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {str(e)}")
            return f"[Error extracting text from PDF: {str(e)}]"
    
    async def _extract_text_from_docx(self, file_path: str) -> str:
        """DOCX'den metin çıkarır"""
        try:
            text = docx2txt.process(file_path)
            return text
        except Exception as e:
            logger.error(f"Error extracting text from DOCX: {str(e)}")
            return f"[Error extracting text from DOCX: {str(e)}]"
    
    async def _extract_text_from_text(self, file_path: str) -> str:
        """Düz metin dosyasından metin çıkarır"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error extracting text from text file: {str(e)}")
            return f"[Error extracting text from text file: {str(e)}]"
    
    async def _extract_text_from_html(self, file_path: str) -> str:
        """HTML'den metin çıkarır"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                html_content = file.read()
            
            if HAS_HTML_PARSER:
                soup = BeautifulSoup(html_content, 'html.parser')
                # Script ve stil etiketlerini kaldır
                for script in soup(["script", "style"]):
                    script.extract()
                
                text = soup.get_text(separator='\n', strip=True)
                
                # Fazla boşlukları temizle
                lines = (line.strip() for line in text.splitlines())
                chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
                text = '\n'.join(chunk for chunk in chunks if chunk)
                
                return text
            else:
                # BeautifulSoup yoksa basit bir temizleme yap
                # HTML etiketlerini kaldır
                clean_text = re.sub(r'<[^>]*>', '', html_content)
                # Fazla boşlukları temizle
                clean_text = re.sub(r'\s+', ' ', clean_text).strip()
                return clean_text
                
        except Exception as e:
            logger.error(f"Error extracting text from HTML: {str(e)}")
            return f"[Error extracting text from HTML: {str(e)}]"
    
    async def _extract_text_from_excel(self, file_path: str) -> str:
        """Excel dosyasından metin çıkarır"""
        if not HAS_EXCEL:
            return "[Excel support not available. Install openpyxl package.]"
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"\n--- Sheet: {sheet_name} ---\n")
                
                for row in sheet.iter_rows(values_only=True):
                    # Boş olmayan hücreleri birleştir
                    row_text = ' | '.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_content.append(row_text)
            
            return '\n'.join(text_content)
            
        except Exception as e:
            logger.error(f"Error extracting text from Excel: {str(e)}")
            return f"[Error extracting text from Excel: {str(e)}]"
    
    async def _extract_text_from_image(self, file_path: str) -> str:
        """Görüntüden OCR ile metin çıkarır"""
        if not self.has_ocr:
            return "[OCR support not available. Install pytesseract and Pillow packages.]"
        
        try:
            # Görüntüyü aç
            with Image.open(file_path) as img:
                # OCR uygula
                text = pytesseract.image_to_string(img)
                return text.strip()
                
        except Exception as e:
            logger.error(f"Error extracting text with OCR: {str(e)}")
            return f"[Error extracting text with OCR: {str(e)}]"
    
    async def _apply_ocr_to_pdf_page(self, pdf_path: str, page_num: int) -> str:
        """PDF sayfasını görüntüye dönüştürür ve OCR uygular"""
        if not self.has_ocr:
            return ""
        
        try:
            # Geçici dosya dizini
            with tempfile.TemporaryDirectory() as temp_dir:
                # PDF sayfasını görüntüye dönüştür (poppler-utils gereklidir)
                image_path = os.path.join(temp_dir, f"page_{page_num}.png")
                os.system(f"pdftoppm -png -f {page_num + 1} -l {page_num + 1} -r 300 {pdf_path} {os.path.join(temp_dir, 'page')}")
                
                # Oluşturulan görüntü dosyasını bul
                image_files = [f for f in os.listdir(temp_dir) if f.startswith(f"page-{page_num + 1}") and f.endswith(".png")]
                
                if not image_files:
                    return ""
                
                image_path = os.path.join(temp_dir, image_files[0])
                
                # OCR uygula
                with Image.open(image_path) as img:
                    text = pytesseract.image_to_string(img)
                    return text.strip()
                    
        except Exception as e:
            logger.error(f"Error applying OCR to PDF page: {str(e)}")
            return ""