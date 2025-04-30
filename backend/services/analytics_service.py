# Last reviewed: 2025-04-30 05:13:15 UTC (User: TeeksssAnalitik)
import logging
from typing import Dict, Any, List, Optional, Union, Tuple
from datetime import datetime, timedelta, timezone
import json
import pandas as pd
import numpy as np
from sqlalchemy import text, func, desc, and_, or_, select, join, case, Integer
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession
import io
import csv
from openpyxl import Workbook

from ..models.user import User
from ..models.document import Document
from ..models.audit import AuditLog
from ..models.query import Query, QuerySource
from ..schemas.analytics import (
    AnalyticsTimeRange, 
    UserActivityReport, 
    DocumentAnalytics,
    QueryAnalytics,
    SystemUsageReport,
    AnalyticsExportFormat
)
from ..services.audit_service import AuditLogType
from ..repositories.user_repository import UserRepository
from ..repositories.document_repository import DocumentRepository
from ..repositories.query_repository import QueryRepository

logger = logging.getLogger(__name__)

class AnalyticsService:
    """
    Raporlama ve analitik servisi
    
    Bu servis şunları sağlar:
    - Kullanıcı etkinlik analizleri
    - Belge kullanım istatistikleri
    - Sistem performans metrikleri
    - Sorgu başarı oranları ve istatistikleri
    - Verilerin CSV/Excel/JSON formatlarında dışa aktarımı
    """
    
    def __init__(self):
        """Analytics servisi başlat"""
        self.user_repository = UserRepository()
        self.document_repository = DocumentRepository()
        self.query_repository = QueryRepository()
        
        # Analiz ayarları
        self.default_limit = 100
        self.default_time_range = AnalyticsTimeRange.LAST_30_DAYS
        
        # Zaman aralığı eşleştirmeleri
        self.time_range_mapping = {
            AnalyticsTimeRange.TODAY: self._get_today_range,
            AnalyticsTimeRange.YESTERDAY: self._get_yesterday_range,
            AnalyticsTimeRange.LAST_7_DAYS: self._get_last_n_days_range(7),
            AnalyticsTimeRange.LAST_30_DAYS: self._get_last_n_days_range(30),
            AnalyticsTimeRange.THIS_MONTH: self._get_this_month_range,
            AnalyticsTimeRange.LAST_MONTH: self._get_last_month_range,
            AnalyticsTimeRange.THIS_QUARTER: self._get_this_quarter_range,
            AnalyticsTimeRange.LAST_QUARTER: self._get_last_quarter_range,
            AnalyticsTimeRange.THIS_YEAR: self._get_this_year_range
        }
    
    def _get_today_range(self) -> Tuple[datetime, datetime]:
        """Bugün için tarih aralığı"""
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)
        return today, tomorrow
    
    def _get_yesterday_range(self) -> Tuple[datetime, datetime]:
        """Dün için tarih aralığı"""
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)
        return yesterday, today
    
    def _get_last_n_days_range(self, days: int) -> callable:
        """Son N gün için tarih aralığı fonksiyonu oluşturma"""
        def get_range() -> Tuple[datetime, datetime]:
            end_date = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59, microsecond=999999)
            start_date = (end_date - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)
            return start_date, end_date
        return get_range
    
    def _get_this_month_range(self) -> Tuple[datetime, datetime]:
        """Bu ay için tarih aralığı"""
        now = datetime.now(timezone.utc)
        start_date = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
        if now.month == 12:
            end_date = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(now.year, now.month + 1, 1, tzinfo=timezone.utc)
        return start_date, end_date
    
    def _get_last_month_range(self) -> Tuple[datetime, datetime]:
        """Geçen ay için tarih aralığı"""
        now = datetime.now(timezone.utc)
        if now.month == 1:
            start_date = datetime(now.year - 1, 12, 1, tzinfo=timezone.utc)
            end_date = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
        else:
            start_date = datetime(now.year, now.month - 1, 1, tzinfo=timezone.utc)
            end_date = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
        return start_date, end_date
    
    def _get_this_quarter_range(self) -> Tuple[datetime, datetime]:
        """Bu çeyrek için tarih aralığı"""
        now = datetime.now(timezone.utc)
        current_quarter = (now.month - 1) // 3 + 1
        start_month = 3 * (current_quarter - 1) + 1
        start_date = datetime(now.year, start_month, 1, tzinfo=timezone.utc)
        
        if current_quarter == 4:
            end_date = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(now.year, start_month + 3, 1, tzinfo=timezone.utc)
        
        return start_date, end_date
    
    def _get_last_quarter_range(self) -> Tuple[datetime, datetime]:
        """Geçen çeyrek için tarih aralığı"""
        now = datetime.now(timezone.utc)
        current_quarter = (now.month - 1) // 3 + 1
        
        if current_quarter == 1:
            last_quarter = 4
            year = now.year - 1
        else:
            last_quarter = current_quarter - 1
            year = now.year
        
        start_month = 3 * (last_quarter - 1) + 1
        start_date = datetime(year, start_month, 1, tzinfo=timezone.utc)
        
        if last_quarter == 4:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, start_month + 3, 1, tzinfo=timezone.utc)
        
        return start_date, end_date
    
    def _get_this_year_range(self) -> Tuple[datetime, datetime]:
        """Bu yıl için tarih aralığı"""
        now = datetime.now(timezone.utc)
        start_date = datetime(now.year, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
        return start_date, end_date
    
    def _get_date_range(self, time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]]) -> Tuple[datetime, datetime]:
        """
        Belirtilen zaman aralığı için başlangıç ve bitiş tarihlerini al
        """
        if isinstance(time_range, tuple):
            return time_range
        
        range_func = self.time_range_mapping.get(time_range, self._get_last_n_days_range(30))
        return range_func()
    
    async def get_user_activity_report(
        self,
        db: AsyncSession,
        organization_id: Optional[str] = None,
        time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]] = None,
        top_users_limit: int = 10
    ) -> UserActivityReport:
        """
        Kullanıcı etkinlik raporu oluştur
        
        Args:
            db: Veritabanı bağlantısı
            organization_id: Organizasyon kimliği filtresi (opsiyonel)
            time_range: Zaman aralığı filtresi
            top_users_limit: En aktif kullanıcı sayısı limiti
            
        Returns:
            UserActivityReport: Kullanıcı etkinlik raporu
        """
        try:
            # Tarih aralığını belirle
            start_date, end_date = self._get_date_range(time_range or self.default_time_range)
            
            # Ana sorgu koşulları
            conditions = [AuditLog.created_at.between(start_date, end_date)]
            if organization_id:
                conditions.append(User.organization_id == organization_id)
            
            # Temel metrikleri al
            total_logins_stmt = (
                select(func.count())
                .select_from(AuditLog)
                .filter(
                    AuditLog.event_type == AuditLogType.AUTH,
                    AuditLog.action == "login",
                    AuditLog.status == "success",
                    AuditLog.created_at.between(start_date, end_date)
                )
            )
            
            if organization_id:
                total_logins_stmt = total_logins_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            total_logins_result = await db.execute(total_logins_stmt)
            total_logins = total_logins_result.scalar() or 0
            
            # Aktif kullanıcı sayısı
            active_users_stmt = (
                select(func.count(func.distinct(AuditLog.user_id)))
                .select_from(AuditLog)
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
            )
            
            if organization_id:
                active_users_stmt = active_users_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            active_users_result = await db.execute(active_users_stmt)
            active_users_count = active_users_result.scalar() or 0
            
            # Toplam kayıtlı kullanıcı sayısı
            total_users_stmt = select(func.count()).select_from(User)
            if organization_id:
                total_users_stmt = total_users_stmt.filter(User.organization_id == organization_id)
            
            total_users_result = await db.execute(total_users_stmt)
            total_users_count = total_users_result.scalar() or 0
            
            # En aktif kullanıcılar (event sayısına göre)
            top_users_stmt = (
                select(
                    AuditLog.user_id,
                    User.email,
                    User.username,
                    User.full_name,
                    func.count().label('event_count')
                )
                .join(User, AuditLog.user_id == User.id)
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
                .group_by(
                    AuditLog.user_id,
                    User.email,
                    User.username,
                    User.full_name
                )
                .order_by(desc('event_count'))
                .limit(top_users_limit)
            )
            
            if organization_id:
                top_users_stmt = top_users_stmt.filter(User.organization_id == organization_id)
            
            top_users_result = await db.execute(top_users_stmt)
            top_users = []
            
            for user_id, email, username, full_name, event_count in top_users_result.all():
                top_users.append({
                    "user_id": str(user_id),
                    "email": email,
                    "username": username,
                    "full_name": full_name,
                    "event_count": event_count
                })
            
            # Günlük etkinlik dağılımı
            daily_activity_stmt = (
                select(
                    func.date_trunc('day', AuditLog.created_at).label('day'),
                    func.count().label('event_count')
                )
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
                .group_by('day')
                .order_by('day')
            )
            
            if organization_id:
                daily_activity_stmt = daily_activity_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            daily_activity_result = await db.execute(daily_activity_stmt)
            daily_activity = []
            
            for day, count in daily_activity_result.all():
                daily_activity.append({
                    "date": day.isoformat(),
                    "count": count
                })
            
            # Etkinlik türü dağılımı
            event_types_stmt = (
                select(
                    AuditLog.event_type,
                    func.count().label('count')
                )
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
                .group_by(AuditLog.event_type)
                .order_by(desc('count'))
            )
            
            if organization_id:
                event_types_stmt = event_types_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            event_types_result = await db.execute(event_types_stmt)
            event_types = []
            
            for event_type, count in event_types_result.all():
                event_types.append({
                    "type": event_type,
                    "count": count
                })
            
            return {
                "total_logins": total_logins,
                "active_users_count": active_users_count,
                "total_users_count": total_users_count,
                "active_users_percentage": round((active_users_count / total_users_count * 100) if total_users_count > 0 else 0, 2),
                "top_active_users": top_users,
                "daily_activity": daily_activity,
                "event_types_distribution": event_types,
                "time_range": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating user activity report: {str(e)}")
            raise
    
    async def get_document_analytics(
        self,
        db: AsyncSession,
        organization_id: Optional[str] = None,
        time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]] = None,
        top_documents_limit: int = 10
    ) -> DocumentAnalytics:
        """
        Belge analitikleri raporu oluştur
        
        Args:
            db: Veritabanı bağlantısı
            organization_id: Organizasyon kimliği filtresi (opsiyonel)
            time_range: Zaman aralığı filtresi
            top_documents_limit: En çok kullanılan belge sayısı limiti
            
        Returns:
            DocumentAnalytics: Belge analitik raporu
        """
        try:
            # Tarih aralığını belirle
            start_date, end_date = self._get_date_range(time_range or self.default_time_range)
            
            # Toplam belge sayısı
            total_docs_stmt = select(func.count()).select_from(Document)
            if organization_id:
                total_docs_stmt = total_docs_stmt.filter(Document.organization_id == organization_id)
            
            total_docs_result = await db.execute(total_docs_stmt)
            total_documents = total_docs_result.scalar() or 0
            
            # Belirtilen dönemde eklenen belge sayısı
            new_docs_stmt = (
                select(func.count())
                .select_from(Document)
                .filter(Document.created_at.between(start_date, end_date))
            )
            
            if organization_id:
                new_docs_stmt = new_docs_stmt.filter(Document.organization_id == organization_id)
            
            new_docs_result = await db.execute(new_docs_stmt)
            new_documents = new_docs_result.scalar() or 0
            
            # En çok sorgulanan belgeler
            top_docs_stmt = (
                select(
                    QuerySource.document_id,
                    Document.title,
                    func.count().label('query_count')
                )
                .join(Document, QuerySource.document_id == Document.id)
                .filter(
                    QuerySource.created_at.between(start_date, end_date)
                )
                .group_by(
                    QuerySource.document_id,
                    Document.title
                )
                .order_by(desc('query_count'))
                .limit(top_documents_limit)
            )
            
            if organization_id:
                top_docs_stmt = top_docs_stmt.filter(Document.organization_id == organization_id)
            
            top_docs_result = await db.execute(top_docs_stmt)
            top_queried_documents = []
            
            for doc_id, title, query_count in top_docs_result.all():
                top_queried_documents.append({
                    "document_id": str(doc_id),
                    "title": title,
                    "query_count": query_count
                })
            
            # Belge boyutu dağılımı
            size_bins = [
                (0, 1024, '0-1 KB'),
                (1025, 10240, '1-10 KB'),
                (10241, 102400, '10-100 KB'),
                (102401, 1048576, '100 KB-1 MB'),
                (1048577, 10485760, '1-10 MB'),
                (10485761, 104857600, '10-100 MB'),
                (104857601, None, '>100 MB')
            ]
            
            size_distribution = []
            for start, end, label in size_bins:
                size_stmt = select(func.count()).select_from(Document)
                
                if start is not None:
                    size_stmt = size_stmt.filter(Document.file_size >= start)
                
                if end is not None:
                    size_stmt = size_stmt.filter(Document.file_size < end)
                
                if organization_id:
                    size_stmt = size_stmt.filter(Document.organization_id == organization_id)
                
                size_result = await db.execute(size_stmt)
                count = size_result.scalar() or 0
                
                size_distribution.append({
                    "range": label,
                    "count": count
                })
            
            # Dosya türleri dağılımı
            file_types_stmt = (
                select(
                    Document.file_type,
                    func.count().label('count')
                )
                .group_by(Document.file_type)
                .order_by(desc('count'))
            )
            
            if organization_id:
                file_types_stmt = file_types_stmt.filter(Document.organization_id == organization_id)
            
            file_types_result = await db.execute(file_types_stmt)
            file_types = []
            
            for file_type, count in file_types_result.all():
                file_types.append({
                    "type": file_type or "unknown",
                    "count": count
                })
            
            # Günlük belge yükleme dağılımı
            daily_uploads_stmt = (
                select(
                    func.date_trunc('day', Document.created_at).label('day'),
                    func.count().label('count')
                )
                .filter(
                    Document.created_at.between(start_date, end_date)
                )
                .group_by('day')
                .order_by('day')
            )
            
            if organization_id:
                daily_uploads_stmt = daily_uploads_stmt.filter(Document.organization_id == organization_id)
            
            daily_uploads_result = await db.execute(daily_uploads_stmt)
            daily_uploads = []
            
            for day, count in daily_uploads_result.all():
                daily_uploads.append({
                    "date": day.isoformat(),
                    "count": count
                })
            
            return {
                "total_documents": total_documents,
                "new_documents": new_documents,
                "documents_growth_percentage": round((new_documents / total_documents * 100) if total_documents > 0 else 0, 2),
                "top_queried_documents": top_queried_documents,
                "file_size_distribution": size_distribution,
                "file_type_distribution": file_types,
                "daily_document_uploads": daily_uploads,
                "time_range": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating document analytics: {str(e)}")
            raise
    
    async def get_query_analytics(
        self,
        db: AsyncSession,
        organization_id: Optional[str] = None,
        time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]] = None
    ) -> QueryAnalytics:
        """
        Sorgu analitikleri raporu oluştur
        
        Args:
            db: Veritabanı bağlantısı
            organization_id: Organizasyon kimliği filtresi (opsiyonel)
            time_range: Zaman aralığı filtresi
            
        Returns:
            QueryAnalytics: Sorgu analitik raporu
        """
        try:
            # Tarih aralığını belirle
            start_date, end_date = self._get_date_range(time_range or self.default_time_range)
            
            # Toplam sorgu sayısı
            total_queries_stmt = (
                select(func.count())
                .select_from(Query)
                .filter(Query.created_at.between(start_date, end_date))
            )
            
            if organization_id:
                total_queries_stmt = total_queries_stmt.filter(Query.organization_id == organization_id)
            
            total_queries_result = await db.execute(total_queries_stmt)
            total_queries = total_queries_result.scalar() or 0
            
            # Günlük sorgu dağılımı
            daily_queries_stmt = (
                select(
                    func.date_trunc('day', Query.created_at).label('day'),
                    func.count().label('count')
                )
                .filter(
                    Query.created_at.between(start_date, end_date)
                )
                .group_by('day')
                .order_by('day')
            )
            
            if organization_id:
                daily_queries_stmt = daily_queries_stmt.filter(Query.organization_id == organization_id)
            
            daily_queries_result = await db.execute(daily_queries_stmt)
            daily_queries = []
            
            for day, count in daily_queries_result.all():
                daily_queries.append({
                    "date": day.isoformat(),
                    "count": count
                })
            
            # Ortalama sorgu süresi
            avg_query_time_stmt = (
                select(func.coalesce(func.avg(Query.processing_time_ms), 0))
                .filter(
                    Query.created_at.between(start_date, end_date),
                    Query.processing_time_ms.isnot(None)
                )
            )
            
            if organization_id:
                avg_query_time_stmt = avg_query_time_stmt.filter(Query.organization_id == organization_id)
            
            avg_query_time_result = await db.execute(avg_query_time_stmt)
            avg_query_time_ms = avg_query_time_result.scalar() or 0
            
            # Sorgu tipine göre dağılım
            query_types_stmt = (
                select(
                    Query.search_type,
                    func.count().label('count')
                )
                .filter(
                    Query.created_at.between(start_date, end_date)
                )
                .group_by(Query.search_type)
                .order_by(desc('count'))
            )
            
            if organization_id:
                query_types_stmt = query_types_stmt.filter(Query.organization_id == organization_id)
            
            query_types_result = await db.execute(query_types_stmt)
            query_types = []
            
            for search_type, count in query_types_result.all():
                query_types.append({
                    "type": search_type or "unknown",
                    "count": count
                })
            
            # Kullanıcı başına ortalama sorgu sayısı
            user_query_counts_stmt = (
                select(
                    Query.user_id,
                    func.count().label('query_count')
                )
                .filter(
                    Query.created_at.between(start_date, end_date),
                    Query.user_id.isnot(None)
                )
                .group_by(Query.user_id)
            )
            
            if organization_id:
                user_query_counts_stmt = user_query_counts_stmt.filter(Query.organization_id == organization_id)
            
            user_query_counts_result = await db.execute(user_query_counts_stmt)
            user_query_counts = [row.query_count for row in user_query_counts_result.all()]
            
            if user_query_counts:
                avg_queries_per_user = sum(user_query_counts) / len(user_query_counts)
            else:
                avg_queries_per_user = 0
            
            # En popüler sorgu anahtar kelimeleri
            # Not: Gerçek uygulamada tokenization ve NLP kullanılabilir
            popular_keywords = await self._extract_popular_query_keywords(
                db, start_date, end_date, organization_id, limit=20
            )
            
            return {
                "total_queries": total_queries,
                "average_query_time_ms": round(avg_query_time_ms, 2),
                "daily_query_counts": daily_queries,
                "query_type_distribution": query_types,
                "average_queries_per_user": round(avg_queries_per_user, 2),
                "popular_keywords": popular_keywords,
                "time_range": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating query analytics: {str(e)}")
            raise
    
    async def _extract_popular_query_keywords(
        self,
        db: AsyncSession,
        start_date: datetime,
        end_date: datetime,
        organization_id: Optional[str] = None,
        limit: int = 20
    ) -> List[Dict[str, Any]]:
        """
        Sorgu metinlerinden popüler anahtar kelimeleri çıkart
        """
        try:
            # Tüm sorguları al
            queries_stmt = select(Query.question).filter(
                Query.created_at.between(start_date, end_date)
            )
            
            if organization_id:
                queries_stmt = queries_stmt.filter(Query.organization_id == organization_id)
            
            queries_result = await db.execute(queries_stmt)
            queries = [row[0] for row in queries_result.all() if row[0]]
            
            # Kelime frekans analizi için basit yaklaşım
            # Gerçek uygulamada NLP ve stopword removal daha iyi olacaktır
            stop_words = set([
                'the', 'a', 'an', 'and', 'or', 'but', 'is', 'are', 'was', 
                'were', 'be', 'been', 'being', 'to', 'of', 'for', 'with',
                'about', 'against', 'between', 'into', 'through', 'during',
                'before', 'after', 'above', 'below', 'from', 'up', 'down',
                'in', 'out', 'on', 'off', 'over', 'under', 'again', 'further',
                'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how',
                'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other',
                'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so',
                'than', 'too', 'very', 's', 't', 'can', 'will', 'just', 'don',
                'should', 'now', 'd', 'll', 'm', 'o', 're', 've', 'y', 'ain',
                'aren', 'couldn', 'didn', 'doesn', 'hadn', 'hasn', 'haven',
                'isn', 'ma', 'mightn', 'mustn', 'needn', 'shan', 'shouldn',
                'wasn', 'weren', 'won', 'wouldn', 'what', 'which', 'who', 'whom',
                'this', 'that', 'these', 'those', 'it', 'its', 'get', 'does',
                'do', 'did', 'have', 'has', 'had', 'having', 'i', 'me', 'my',
                'myself', 'we', 'our', 'ours', 'ourselves', 'you', "you're",
                "you've", "you'll", "you'd", 'your', 'yours', 'yourself',
                'yourselves', 'he', 'him', 'his', 'himself', 'she', "she's",
                'her', 'hers', 'herself'
            ])
            
            # Kelime frekansı hesapla
            word_freq = {}
            for query in queries:
                # Noktalama işaretlerini temizle ve küçük harfe çevir
                clean_query = ''.join(c.lower() if c.isalpha() or c.isspace() else ' ' for c in query)
                words = clean_query.split()
                
                for word in words:
                    if len(word) > 2 and word.lower() not in stop_words:
                        word_freq[word.lower()] = word_freq.get(word.lower(), 0) + 1
            
            # En popüler kelimeleri sırala
            keywords = [
                {"keyword": word, "count": count}
                for word, count in sorted(
                    word_freq.items(),
                    key=lambda x: x[1],
                    reverse=True
                )[:limit]
            ]
            
            return keywords
            
        except Exception as e:
            logger.error(f"Error extracting popular keywords: {str(e)}")
            return []
    
    async def get_system_usage_report(
        self,
        db: AsyncSession,
        organization_id: Optional[str] = None,
        time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]] = None
    ) -> SystemUsageReport:
        """
        Sistem kullanım raporu oluştur
        
        Args:
            db: Veritabanı bağlantısı
            organization_id: Organizasyon kimliği filtresi (opsiyonel)
            time_range: Zaman aralığı filtresi
            
        Returns:
            SystemUsageReport: Sistem kullanım raporu
        """
        try:
            # Tarih aralığını belirle
            start_date, end_date = self._get_date_range(time_range or self.default_time_range)
            
            # Başarılı ve başarısız sorgu oranları
            success_failure_stmt = (
                select(
                    Query.has_error,
                    func.count().label('count')
                )
                .filter(
                    Query.created_at.between(start_date, end_date)
                )
                .group_by(Query.has_error)
            )
            
            if organization_id:
                success_failure_stmt = success_failure_stmt.filter(Query.organization_id == organization_id)
            
            success_failure_result = await db.execute(success_failure_stmt)
            
            success_count = 0
            error_count = 0
            
            for has_error, count in success_failure_result.all():
                if has_error:
                    error_count = count
                else:
                    success_count = count
            
            total_count = success_count + error_count
            
            # Gün içi aktivite paternleri (saate göre)
            hourly_activity_stmt = (
                select(
                    func.extract('hour', AuditLog.created_at).cast(Integer).label('hour'),
                    func.count().label('count')
                )
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
                .group_by('hour')
                .order_by('hour')
            )
            
            if organization_id:
                hourly_activity_stmt = hourly_activity_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            hourly_activity_result = await db.execute(hourly_activity_stmt)
            hourly_activity = []
            
            for hour, count in hourly_activity_result.all():
                hourly_activity.append({
                    "hour": hour,
                    "count": count
                })
            
            # Haftalık aktivite paternleri (haftanın günlerine göre)
            weekly_activity_stmt = (
                select(
                    func.extract('dow', AuditLog.created_at).cast(Integer).label('day_of_week'),
                    func.count().label('count')
                )
                .filter(
                    AuditLog.created_at.between(start_date, end_date)
                )
                .group_by('day_of_week')
                .order_by('day_of_week')
            )
            
            if organization_id:
                weekly_activity_stmt = weekly_activity_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            weekly_activity_result = await db.execute(weekly_activity_stmt)
            weekly_activity = []
            
            for day, count in weekly_activity_result.all():
                # PostgreSQL'de 0=Pazar, 1=Pazartesi, ... 6=Cumartesi
                day_name = ['Pazar', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi'][day]
                weekly_activity.append({
                    "day": day,
                    "day_name": day_name,
                    "count": count
                })
            
            # Ortalama ve maksimum eşzamanlı oturum sayısı
            # Not: Bu metrik için gerçek veri yoksa tahmini değerler kullanılabilir
            active_sessions_stmt = (
                select(
                    func.date_trunc('hour', AuditLog.created_at).label('hour'),
                    func.count(func.distinct(AuditLog.user_id)).label('active_users')
                )
                .filter(
                    AuditLog.created_at.between(start_date, end_date),
                    AuditLog.event_type == AuditLogType.AUTH,
                    or_(
                        AuditLog.action == "login",
                        AuditLog.action == "refresh_token"
                    ),
                    AuditLog.status == "success"
                )
                .group_by('hour')
                .order_by('hour')
            )
            
            if organization_id:
                active_sessions_stmt = active_sessions_stmt.join(
                    User, AuditLog.user_id == User.id
                ).filter(
                    User.organization_id == organization_id
                )
            
            active_sessions_result = await db.execute(active_sessions_stmt)
            active_sessions = [row.active_users for row in active_sessions_result.all()]
            
            max_concurrent_sessions = max(active_sessions, default=0)
            avg_concurrent_sessions = sum(active_sessions) / len(active_sessions) if active_sessions else 0
            
            return {
                "total_queries": total_count,
                "successful_queries": success_count,
                "failed_queries": error_count,
                "success_rate_percentage": round((success_count / total_count * 100) if total_count > 0 else 0, 2),
                "hourly_activity_pattern": hourly_activity,
                "weekly_activity_pattern": weekly_activity,
                "max_concurrent_sessions": max_concurrent_sessions,
                "avg_concurrent_sessions": round(avg_concurrent_sessions, 2),
                "time_range": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating system usage report: {str(e)}")
            raise
    
    async def export_analytics_data(
        self,
        db: AsyncSession,
        report_type: str,
        format: AnalyticsExportFormat,
        organization_id: Optional[str] = None,
        time_range: Union[AnalyticsTimeRange, Tuple[datetime, datetime]] = None
    ) -> Tuple[bytes, str]:
        """
        Analitik verilerini dışa aktar
        
        Args:
            db: Veritabanı bağlantısı
            report_type: Rapor tipi ('user_activity', 'document_analytics', 'query_analytics', 'system_usage')
            format: Dışa aktarım formatı (CSV, Excel, JSON)
            organization_id: Organizasyon kimliği filtresi (opsiyonel)
            time_range: Zaman aralığı filtresi
            
        Returns:
            Tuple[bytes, str]: Dosya içeriği ve dosya adı
        """
        try:
            # Rapor tipine göre veriyi al
            data = None
            filename_prefix = "analytics_report"
            
            if report_type == "user_activity":
                data = await self.get_user_activity_report(db, organization_id, time_range)
                filename_prefix = "user_activity_report"
            elif report_type == "document_analytics":
                data = await self.get_document_analytics(db, organization_id, time_range)
                filename_prefix = "document_analytics_report"
            elif report_type == "query_analytics":
                data = await self.get_query_analytics(db, organization_id, time_range)
                filename_prefix = "query_analytics_report"
            elif report_type == "system_usage":
                data = await self.get_system_usage_report(db, organization_id, time_range)
                filename_prefix = "system_usage_report"
            else:
                raise ValueError(f"Unsupported report type: {report_type}")
            
            # Veri formatına göre dönüştür
            if format == AnalyticsExportFormat.JSON:
                file_content = json.dumps(data, indent=2, default=str).encode('utf-8')
                filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                
            elif format == AnalyticsExportFormat.CSV:
                file_content = self._convert_to_csv(data)
                filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                
            elif format == AnalyticsExportFormat.EXCEL:
                file_content = self._convert_to_excel(data)
                filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
            return file_content, filename
            
        except Exception as e:
            logger.error(f"Error exporting analytics data: {str(e)}")
            raise
    
    def _convert_to_csv(self, data: Dict[str, Any]) -> bytes:
        """
        Veriyi CSV formatına dönüştür
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Düz bir veri yapısına dönüştür
        flat_data = self._flatten_data(data)
        
        # Başlık satırı
        writer.writerow(flat_data.keys())
        
        # Veri satırı
        writer.writerow(flat_data.values())
        
        return output.getvalue().encode('utf-8')
    
    def _convert_to_excel(self, data: Dict[str, Any]) -> bytes:
        """
        Veriyi Excel formatına dönüştür
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Düz bir veri yapısına dönüştür
        flat_data = self._flatten_data(data)
        
        # Başlık satırı
        for col_idx, header in enumerate(flat_data.keys(), start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Veri satırı
        for col_idx, value in enumerate(flat_data.values(), start=1):
            ws.cell(row=2, column=col_idx, value=str(value))
        
        # Liste türündeki verileri ayrı sayfalara ekle
        for key, value in data.items():
            if isinstance(value, list) and len(value) > 0:
                list_ws = wb.create_sheet(title=key[:31])  # Excel sayfa adı uzunluk sınırı
                
                # Liste elemanlarının anahtarlarını al
                if value and isinstance(value[0], dict):
                    headers = list(value[0].keys())
                    
                    # Başlık satırı
                    for col_idx, header in enumerate(headers, start=1):
                        list_ws.cell(row=1, column=col_idx, value=header)
                    
                    # Veri satırları
                    for row_idx, item in enumerate(value, start=2):
                        for col_idx, header in enumerate(headers, start=1):
                            list_ws.cell(row=row_idx, column=col_idx, value=str(item.get(header, '')))
        
        # Excel'i bytes olarak kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _flatten_data(self, data: Dict[str, Any], prefix: str = '') -> Dict[str, Any]:
        """
        İç içe veri yapısını düz bir yapıya dönüştür
        """
        result = {}
        
        for key, value in data.items():
            new_key = f"{prefix}_{key}" if prefix else key
            
            if isinstance(value, dict):
                # İç içe sözlükleri düzleştir
                nested_result = self._flatten_data(value, new_key)
                result.update(nested_result)
            elif isinstance(value, list):
                # Listeleri JSON string'e dönüştür
                result[new_key] = json.dumps(value, default=str)
            else:
                # Diğer değerler doğrudan ekle
                result[new_key] = value
        
        return result